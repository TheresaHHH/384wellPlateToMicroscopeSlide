import pandas as pd
import numpy as np
import openpyxl
import sys

fileName= sys.argv[1]
op0 = input('The fileName is '+ fileName +'. Enter 0 to input new fileName:')
if op0 == '0':
    fileName = input("Enter fileName.xlsx:");
else:
    print(fileName + ' will be read. Please close it if you have opened it.');

#####it only runs when when you run this file it more than once 
newSheet= 'RotatedTable';
wb = openpyxl.load_workbook(fileName);
if newSheet in wb.sheetnames:
    del wb[newSheet]
    wb.save(fileName);
  
##Please make sure the extension is .xlsx
#reads the excel file
excel_File = pd.ExcelFile(fileName);    
excelSheets = excel_File.sheet_names;

#How many columns and rows in each small table?
firstTableColumn = 12;
firstTableRow = 4;
firstTables = [];

#Read the excel file and create tables
for sheets in range(len(excelSheets)):
    #Dataframe
    file = excel_File.parse(excelSheets[sheets], header=None, names=None, index_col=None, usecols=None) ;

    ##How many tables are on the left side?
    for j in range(4):
        firstTable=[];
        for i in range(firstTableRow):
            tempRow = file.loc[i+j*firstTableRow, 0 : firstTableColumn-1].values.tolist();
            firstTable.extend(tempRow);
        firstTables.insert(len(firstTables), firstTable);

    ##How many tables are on the right side?
    for j in range(7):
        firstTable=[];
        for i in range(firstTableRow):
            tempRow = file.loc[i+j*firstTableRow, firstTableColumn: firstTableColumn*2-1].values.tolist();
            firstTable.extend(tempRow);
        firstTables.insert(len(firstTables), firstTable);        
    print('sheet: '+ str(sheets) + ' is read.');

'''
#testing    
print('firstTables')
for i in range(len(firstTables)):
    print(str(i))
    print(firstTables[i])
'''


#Organizes the cells into tables based on their coordinators
#Stores three occurrences of the original sequence.
#Each element is moved one position to the right, and the last element wraps around to the beginning.
#Adds blanks to each table
zippedTables0 = [list(x) * 3 for x in zip(*firstTables)]
zippedTables0 = [ [sublist[-1]] + sublist[:-1]+['blank']+['blank']+['blank'] for sublist in zippedTables0];


##Please change 6 to the desired value. How many items in each rows in pre-rotated tables? 
grouped = [[sublist[i:i + 6] for i in range(0, len(sublist), 6)] for sublist in zippedTables0];

'''
#testing
print('grouped')
for i in range(len(grouped)):
    print(grouped[i])
'''

for j in range(len(grouped)):
        temp0 = np.rot90(grouped[j],-1);
        grouped[j] = temp0.tolist();
        zippedTables0[j]=grouped[j] ;

'''
#testing
print('zippedTables0')
for i in range(len(zippedTables0)):
    print(chr(65+int(i/12))+str(int(i%12)+1))
    print(zippedTables0[i])
'''

#48 pre-rotated tables in 4 rows. 12 tables in a row. 
##Please change the values if different. 
secondTableColumn = 12;
secondTableRow = 4;
zippedTables1 = [];

for i in range(secondTableRow):
    temp0 = zippedTables0[i*secondTableColumn:secondTableColumn*(i+1)];
    
    zippedTables1.insert(i, zippedTables0[i*secondTableColumn:secondTableColumn*(i+1)]);
    
zippedTables1.reverse();

'''    
#testing
print('zippedTables1')
for i in range(len(zippedTables1)):
    print(str(i))
    for j in range(len(zippedTables1[i])):
        print(zippedTables1[i][j])
'''

'''
print(str(len(zippedTables1[0][0][0]))) #6
print(zippedTables1[0][0][5])
print(str(len(zippedTables1[0][0]))) #6
print(zippedTables1[0][11])
print(str(len(zippedTables1[0])))
print(zippedTables1[3]) #12
print(str(len(zippedTables1))) #4
'''

##Please change the 72 to the acutal number of rows in rotatedTables.
finalexceltext0 = [[] for _ in range(72)]

Raws=f'{"Block":<13}{"Raw":<11}{"Column":<12}{"GeneID":<25}'+'\n';
 
##Please replace 6 with the desired value. How many cells in a row? 
for i in range(len(zippedTables1)):
    for j in range(len(zippedTables1[0])):
        temp = [zippedTables1[i][j][z : z+6] for z in range(0, len(zippedTables1[i][j]),6)];
        for k in range(len(zippedTables1[0][0])):
            if i!=0 :
                finalexceltext0[j*6+k].extend(zippedTables1[i][j][k]);     
            else:
                finalexceltext0[j*6+k]= zippedTables1[i][j][k]; 
            for z, g in enumerate(zippedTables1[i][j][k]):
                Raws = Raws + f'{str((i)*secondTableColumn+j+1):<13}{str(k+1):<11}{str(z+1):<12}{g:<25}'+'\n';
                             
Raws= Raws+'\n'+'Created by HsuanJang and HsuanHua Hung <3'                             
print(Raws)
#print(finalexceltext0);

#Write a txt file. 
with open('BeautifulTable.txt', 'w') as file:
    file.write(Raws);
print('A txt file is created.')

# Write an excel file    
new_sheet = wb.create_sheet(title = newSheet);
ws = wb[newSheet];
for row_index, row in enumerate(finalexceltext0):
    for col_index, value in enumerate(row):
        ws.cell(row = row_index + 1, column = col_index + 1, value = value)
wb.save(fileName)
print('Sheet RotatedTable is added to '+ fileName);

op1 = input('Do you want to create a new excel file? Enter 0 to create it.')
if op1 == '0':
  ex = pd.DataFrame(finalexceltext0)
  ex.to_excel('BeautifulTables.xlsx', index=False, header=False)
else:
  print('The End.')
  
