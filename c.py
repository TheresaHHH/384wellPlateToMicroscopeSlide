import pandas as pd
import numpy as np
from openpyxl import load_workbook
from  openpyxl.utils.dataframe import dataframe_to_rows

# Load the workbook
fileName='Result.xlsx'
wb = load_workbook(fileName)
sheetName=wb.sheetnames[0]

# Select the Result sheet
sheet = wb[sheetName]
tables1 = []
for j in range(0, 24, 6): #swith these two lines.  col by col/row by row
        for i in range(0, 72, 6):#swith these two lines
            # Extract the 7x7block
            block = [[cell.value for cell in row[j:j+6]] for row in sheet[i+1:i+6]]
            # Append the block to the tables list
            tables1.append(block)
#print(tables1[0])  
#print(tables1[1])          

#####export data into a txt file#######

with open('result.txt', 'w') as f:
    # Write the header line
    f.write(f"{'Block':<5} {'row':<5} {'column':<7} {'data'}\n")
    
    for i in range(1,49):
        for row in range(6):
            for col in range(6):
                data=tables1[i-1][row][col]
                #f.write(f'{i}, {row+1}, {col+1}, {data}\n')
                f.write(f"{i:<5} {row + 1:<5} {col + 1:<7} {data}\n")
                
print("result.txt is generated")
